import csv
import io

from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import get_template
<<<<<<< HEAD


from Glenda_App.models import Menu

from inventory_app.forms import Raw_materials_StockForm, Finished_Goods_StockForm,Finished_Goods_RequestForm
from inventory_app.models import RawMaterialsStock,Finished_Goods_Stock,Finished_Goods_Request
from production_app.models import water_Finished_Goods,water_Finished_goods_category
from register_app.models import department

<<<<<<< HEAD
=======
=======
>>>>>>> 9a913e1cb15a3bc11ff9238e022be96a8748665e
=======
>>>>>>> 64b3d20419b8417ad6be80a226ded5fd35487072
>>>>>>> master
from inventory_app.forms import Raw_materials_StockForm, Finished_Goods_StockForm,Damaged_Goods_StockForm
from inventory_app.models import RawMaterialsStock, Finished_Goods_Stock, Damaged_Goods_Stock
from production_app.models import water_Finished_Goods,damaged_Goods,Damaged_good_category
from purchase_app.models import RawMaterials
from Glenda_App.models import Menu
from xhtml2pdf import pisa
from django.db.models import Q
from openpyxl.styles import Font,Alignment
from openpyxl import Workbook
from django.http import FileResponse
from reportlab.pdfgen import canvas

from django.db.models import Sum

# Create your views here.




def raw_materials_view(request):
    menus = Menu.objects.prefetch_related('submenus').all()
    raw_materials = RawMaterials.objects.prefetch_related('stocks').all()

    # Calculate total stock for each raw material
    total_stocks = {
        material.id: material.stocks.aggregate(total=Sum('stock'))['total'] or 0
        for material in raw_materials
    }

    return render(request, 'inventory/view_raw_materials.html', {'view': raw_materials, 'menus': menus, 'total_stocks': total_stocks})

def update_stocks(request, id):
    menus = Menu.objects.prefetch_related('submenus').all()
    raw_material = get_object_or_404(RawMaterials, id=id)

    if request.method == 'POST':
        form = Raw_materials_StockForm(request.POST)
        if form.is_valid():
            stock_entry = form.save(commit=False)
            stock_entry.raw_materials = raw_material  # Set the raw material
            stock_entry.save()  # Save the new stock entry

            # Update the total stock for the raw material
            total_stock = raw_material.stocks.aggregate(total=Sum('stock'))['total'] or 0
            raw_material.total_stock = total_stock
            raw_material.save()  # Save the updated total stock

            return redirect('Raw_materials_view')  # Redirect to the list view
    else:
        form = Raw_materials_StockForm()

    return render(request, 'inventory/add_stock.html', {
        'form': form,
        'menus': menus,
        'raw_material': raw_material
    })




def update_finished_goods_stocks(request, id):
    menus = Menu.objects.prefetch_related('submenus').all()
    finished_goods = get_object_or_404(water_Finished_Goods, id=id)

    if request.method == 'POST':
        form = Finished_Goods_StockForm(request.POST)
        if form.is_valid():
            stock_entry = form.save(commit=False)
            stock_entry.finished_goods = finished_goods  # Set the raw material
            stock_entry.save()  # Save the new stock entry

            # Update the total stock for the raw material
            total_stock = finished_goods.stocks.aggregate(total=Sum('stock'))['total'] or 0
            finished_goods.total_stock = total_stock
            finished_goods.save()  # Save the updated total stock

            return redirect('finishedgoods_stock_view')  # Redirect to the list view
    else:
        form = Finished_Goods_StockForm()

    return render(request, 'inventory/add_finishedgoods_stock.html', {
        'form': form,
        'menus': menus,
        'finished_goods': finished_goods
    })


def finishedgoods_stock_view(request):
    menus = Menu.objects.prefetch_related('submenus').all()
    finished_goods = water_Finished_Goods.objects.prefetch_related('stocks').all()
    categories = water_Finished_goods_category.objects.all()

    # Calculate total stock for each raw material
    total_stocks = {
        material.id: material.stocks.aggregate(total=Sum('stock'))['total'] or 0
        for material in finished_goods
    }

    return render(request, 'inventory/view_finished_goods.html', {'view': finished_goods, 'menus': menus, 'total_stocks': total_stocks,'categories':categories})


def finishedgoods_stock_history(request,id):
    menus = Menu.objects.prefetch_related('submenus').all()
    finished_good = Finished_Goods_Stock.objects.filter(finished_goods_id=id)
    print(finished_good)


    return render(request,'inventory/view_finished_goods_history.html',{'data':finished_good,'menus': menus})

def update_damaged_goods_stocks(request, id):
    menus = Menu.objects.prefetch_related('submenus').all()
    damaged_goods = get_object_or_404(damaged_Goods, id=id)

    if request.method == 'POST':
        form = Damaged_Goods_StockForm(request.POST)
        if form.is_valid():
            stock_entry = form.save(commit=False)
            stock_entry.damaged = damaged_goods  # Set the raw material
            stock_entry.save()  # Save the new stock entry

            # Update the total stock for the raw material
            total_stock = damaged_goods.stocks.aggregate(total=Sum('stock'))['total'] or 0
            damaged_goods.total_stock = total_stock
            damaged_goods.save()  # Save the updated total stock

            return redirect('damagedgoods_stock_view')  # Redirect to the list view
    else:
        form = Damaged_Goods_StockForm()

    return render(request, 'inventory/add_damage_stock.html', {
        'form': form,
        'menus': menus,
        'damaged_goods': damaged_goods
    })


def damagedgoods_stock_view(request):
    menus = Menu.objects.prefetch_related('submenus').all()
    damaged_goods = damaged_Goods.objects.prefetch_related('stocks').all()

    # Calculate total stock for each raw material
    total_stocks = {
        material.id: material.stocks.aggregate(total=Sum('stock'))['total'] or 0
        for material in damaged_goods
    }

    return render(request, 'inventory/view_damaged_goods.html', {'view': damaged_goods, 'menus': menus, 'total_stocks': total_stocks})


def finishedgoods_stock_history(request,id):
    menus = Menu.objects.prefetch_related('submenus').all()
    finished_good = Finished_Goods_Stock.objects.filter(finished_goods_id=id)
    print(finished_good)
    context = {
        'data': finished_good,
        'menus': menus,
        'id': id
    }


    return render(request,'inventory/view_finished_goods_history.html',context)

def finishedgoods_stock_history_pdf(request, id):
    view = Finished_Goods_Stock.objects.filter(finished_goods_id=id).first()
    filename = f"{view.finished_goods.category.category_name}_inventory_report.pdf"

    finished_good = Finished_Goods_Stock.objects.filter(finished_goods_id=id)
    context = {'view': finished_good, 'name':view}

    template = get_template('inventory/finishedgoods_stock_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pdf_status = pisa.CreatePDF(
        html,
        dest=response
    )
    return response



def finishedgoods_stock_history_csv(request,id):

    data = Finished_Goods_Stock.objects.filter(finished_goods_id=id).first()
    filename = f"{data.finished_goods.category.category_name}_inventory_report.csv"

    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Create a CSV writer object
    writer = csv.writer(response)

    # Write the header row (similar to table headers)
    writer.writerow(['Category', 'Name', 'Size', 'Stock', 'Date'])

    # Fetch data from the database (replace Inventory with your model)
    finished_good = Finished_Goods_Stock.objects.filter(finished_goods_id=id)

    # Iterate through the data and write each row to the CSV
    for i in finished_good:
        writer.writerow([
            i.finished_goods.category.category_name,
            i.finished_goods.name,
            i.finished_goods.size,
            i.stock,
            i.date.strftime('%Y-%m-%d')  # Formatting date as needed
        ])

    return response



def raw_materials_stock_history(request,id):
    menus = Menu.objects.prefetch_related('submenus').all()
    raw_materials = RawMaterialsStock.objects.filter(raw_materials_id=id)
    context = {
        'data': raw_materials,
        'menus': menus,
        'id': id
    }

    return render(request,'inventory/view_raw_materials_history.html',context)


def raw_materials_stock_pdf(request, id):
    menus = Menu.objects.prefetch_related('submenus').all()
    view = RawMaterialsStock.objects.filter(raw_materials_id=id).first()
    data = RawMaterialsStock.objects.filter(raw_materials_id=id)
    context = {'view': data, 'menu': menus, 'name':view}

    template = get_template('inventory/raw_material_stock_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{view.raw_materials.name}_raw_material.pdf"'

    pdf_status = pisa.CreatePDF(
        html,
        dest=response
    )
    return response



def damagedgoods_stock_history(request,id):
    menus = Menu.objects.prefetch_related('submenus').all()
    damaged_goods = Damaged_Goods_Stock.objects.filter(damaged_id=id)
    print("success")

    return render(request,'inventory/damaged_stock_history.html',{'menus':menus,'data':damaged_goods,'id':id})


def generate_pdf(request,id):
    view = Damaged_Goods_Stock.objects.filter(damaged_id=id).first()
    filename = f"{view.damaged.category.category_name}_inventory_report.pdf"

    damaged = Damaged_Goods_Stock.objects.filter(damaged_id=id)
    context = {'view': damaged, 'name': view}

    template = get_template('inventory/damaged_stock_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pdf_status = pisa.CreatePDF(
        html,
        dest=response
    )
    return response

def generate_csv(request, id):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="damaged_goods_stock_list.csv"'

    # Create a CSV writer
    writer = csv.writer(response)

    # Write the header row (this matches the columns in your HTML table)
    writer.writerow(['Category', 'Name', 'Stock', 'Description'])

    # Retrieve all items from the database
    items = Damaged_Goods_Stock.objects.filter(damaged_id=id)

    # Write data rows
    for item in items:
        writer.writerow([item.damaged.category, item.damaged.name, item.stock, item.damaged.description])

    # Return the CSV file as a response
    return response
<<<<<<< HEAD
=======
def generate_full_pdf(request):
    view = Damaged_Goods_Stock.objects.all()
    filename = "full_analysis_report.pdf"

    damaged = Damaged_Goods_Stock.objects.all()
    context = {'view': damaged, 'name': view}

    template = get_template('inventory/damaged_full_stock_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pdf_status = pisa.CreatePDF(
        html,
        dest=response
    )
    return response

def damaged_search(request):
    categories = Damaged_good_category.objects.all()
    damaged_goods_list = damaged_Goods.objects.all()
    menus = Menu.objects.prefetch_related('submenus').all()

    context = {
        'view': damaged_goods_list,  # This is the queryset for the table in the HTML
        'categories': categories,
        'menus':menus
    }

    # Handle search requests
    if request.method == 'GET':
        damaged_name = request.GET.get('name', None)
        damaged_category = request.GET.get('category', None)

        # Build filters
        filters = Q()

        # If the search name is provided, filter by the name in damaged_Goods
        if damaged_name:
            filters &= Q(name__icontains=damaged_name)

        # If a valid category is selected, filter by the category
        if damaged_category:
            filters &= Q(category_id=damaged_category)  # Use category_id to filter

        # Apply the filters to the queryset if filters are provided
        if filters:
            damaged_goods_list = damaged_Goods.objects.filter(filters)

        # Update the context with the filtered queryset
        context['view'] = damaged_goods_list

    return render(request, 'inventory/view_damaged_goods.html', context)
>>>>>>> master


def generate_excel(request):
    # Create a Workbook
    wb = Workbook()
    ws = wb.active

    # Write a headline row that spans all columns
    ws.merge_cells('A1:D1')  # Adjust based on your number of columns
    ws['A1'] = 'Damaged Goods Stock List'

    # Set font properties including bold and increased font size
    headline_font = Font(bold=True, size=14)  # Set bold and font size to 14
    ws['A1'].font = headline_font  # Apply the font to the headline

    # Center the headline text
    ws['A1'].alignment = Alignment(horizontal='center')  # Center alignment

    # Write the header row
    ws.append([ 'Category', 'Name', 'Total Stock'])

    # Retrieve all items from the database
    items = Damaged_Goods_Stock.objects.all()

    # Write data rows
    for item in items:
        ws.append([
            item.damaged.category.category_name,  # Category name
            item.damaged.name,  # Item name
            item.stock  # Total stock
        ])

    # Create an HTTP response with the appropriate Excel header
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="damaged_goods_stock_list.xlsx"'

    # Save the workbook to the response
    wb.save(response)
<<<<<<< HEAD
    return response

def search(request):
    items = []
    if request.method == "POST":
        search_obj = request.POST.get("search")
        if search_obj:
            items = damaged_Goods.objects.filter(
                Q(description__icontains=search_obj) | Q(name__icontains=search_obj)
            )
        else:
            data = "No search item"
    return render(request, 'inventory/searched_damaged_goods.html', {"items": items})
<<<<<<< HEAD



def finishedgoods_search(request):
    menus = Menu.objects.prefetch_related('submenus').all()

    finished_goods_list = water_Finished_Goods.objects.all()
    categories = water_Finished_goods_category.objects.all()

    context = {
        'view': finished_goods_list,  # This is the queryset for the table in the HTML
        'categories': categories,  # Pass the categories for the dropdown
        'menus': menus
    }

    # Handle search requests
    if request.method == 'POST':
        search_name = request.POST.get('name', None)
        search_category = request.POST.get('category', None)

        # Build filters
        filters = Q()

        # If the search name is provided, filter by name (case insensitive)
        if search_name:
            filters &= Q(name__icontains=search_name)

        # If a valid category is selected, filter by category
        if search_category and search_category.isdigit():
            filters &= Q(category_id=int(search_category))

        # Apply the filters to the queryset
        if filters:
            finished_goods_list = water_Finished_Goods.objects.filter(filters)

        # Update the context with the filtered queryset
        context['view'] = finished_goods_list

    return render(request, 'inventory/view_finished_goods.html', context)


def finishedgoods_message_request(request):
    menus = Menu.objects.prefetch_related('submenus').all()

    dept = department.objects.all()
    category = water_Finished_goods_category.objects.all()
    name = water_Finished_Goods.objects.all()
    view = Finished_Goods_Request.objects.all()

    if request.method == 'POST':
        form = Finished_Goods_RequestForm(request.POST)
        print(request.POST)  # Debug: Print the form data
        if form.is_valid():
            print("Form is valid")  # Debug: Form is valid
            form_entry = form.save(commit=False)
            # Ensure foreign key is assigned correctly
            form_entry.name = form.cleaned_data['name']
            form_entry.status = 'Pending'
            form_entry.save()
            return redirect('Raw_materials_view')
        else:
            print(form.errors)  # Debug: Print form errors
    else:
        form = Finished_Goods_RequestForm()

    return render(request, 'inventory/finishedgoods_message_request.html', {
        'form': form,
        'menus': menus,
        'department': dept,
        'category': category,
        'name': name,
        'view': view
    })




=======
>>>>>>> 9a913e1cb15a3bc11ff9238e022be96a8748665e
=======
    return response
>>>>>>> 64b3d20419b8417ad6be80a226ded5fd35487072
>>>>>>> master
